from datetime import datetime
from typing import Any

from django import forms
from django.contrib import admin, messages
from django.db import models, transaction
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import path
from django.utils.datastructures import MultiValueDict
from openpyxl import Workbook, load_workbook

# from inventory.models import (
#     ColorVariation,
#     ProductColorVariationImage,
#     ProductTextVariation,
#     ProductVariation,
#     TextVariation,
#     Variation,
# )


class XlsxImportForm(forms.Form):
    xlsx_file = forms.FileField(widget=forms.FileInput(attrs={"accept": ".xlsx"}))


class ImportableExportableAdmin(admin.ModelAdmin):
    """
    This is a base class for admins supporting import and export from and to
    xlsx files of their data.
    Note that only code that is relevant to *all* subclasses should live here.
    You can override functions as necessary for model-specific logic
    """

    import_form: forms.Form = XlsxImportForm
    import_related_fields: set[str] = set()
    import_multiple_value_splitter: str = "|||"
    export_fields: set[str] = set()
    import_excluded_fields: set[str] = set()
    import_excluded_fields_indexes: list = []

    change_list_template = "admin/import_changelist.html"

    def get_urls(self):
        """
        Adding a url to import data file in the existing urls.
        """
        urls = super().get_urls()
        my_urls = [
            path("import-data/", self.import_xlsx),
        ]
        return my_urls + urls

    def import_xlsx(self, request, **kwargs):
        """
        In the GET request, showing a form to submit a xlsx file
        and in POST request, creating the records in the model.
        """
        model_name = self.model._meta.model_name

        if request.method == "POST":
            try:
                # use an atomic transaction so that no data is saved if any
                # single object failed to save
                with transaction.atomic():
                    workbook = load_workbook(
                        request.FILES.pop("xlsx_file")[0],
                        data_only=True,
                        read_only=True,
                    )
                    created, updated = self.import_parse_and_save_xlsx_data(
                        kwargs, workbook
                    )

                    self.message_user(
                        request,
                        (
                            f"Your data file has been imported - {created} "
                            f"{model_name}(s) were created and {updated} "
                            f"{model_name}(s) were updated"
                        ),
                    )
            except RecordImportError as ex:
                for error, row_numbers in ex.errors.items():
                    self.message_user(
                        request,
                        (
                            f"Row{'s' if len(row_numbers) > 1 else ''} "
                            f"{', '.join(str(r) for r in row_numbers)} failed "
                            f"with error: {error}"
                        ),
                        level=messages.ERROR,
                    )
                return redirect(request.path)
            except Exception as ex:
                self.message_user(
                    request,
                    "An unknown error occurred: {}".format(ex),
                    level=messages.ERROR,
                )
                return redirect(request.path)

            return redirect(
                f"admin:{self.model._meta.app_label}_{model_name}_changelist"
            )

        form = self.import_form()
        payload = {"form": form, "model": model_name.capitalize()}
        return render(request, "admin/import_form.html", payload)

    def import_parse_and_save_xlsx_data(
        self, extra_params: dict[str, Any], workbook: Workbook
    ) -> tuple[int, int]:
        raise NotImplementedError()

    def import_parse_field(
        self,
        name: str,
        value: str,
    ):
        field_type = self.model._meta.get_field(name).get_internal_type()

        if field_type in ("CharField", "TextField"):
            return value
        elif field_type == "BooleanField":
            if isinstance(value, bool):
                return value
            else:
                return str(value).lower() in ("true", "yes", "1") if value else None
        elif field_type in ("IntegerField", "BigAutoField"):
            return int(value) if value else None
        elif field_type == "DateField":
            if isinstance(value, datetime):
                return value
            else:
                return datetime.fromisoformat(value).date() if value else None
        elif field_type == "FloatField":
            return float(value) if value else None
        elif field_type == "FileField":
            return value

        raise ValueError(f"Failed to parse unknown field type: {field_type}")

    def import_parse_related_field(
        self,
        name: str,
        value: str,
        extra_files: MultiValueDict,
        main_record: models.Model,
    ):
        raise NotImplementedError(
            "`import_parse_related_field` must be implemented if "
            "`import_related_fields` are supplied to create related objects "
            "properly"
        )

    def _import_split_field_value(self, value: str):
        if not value:
            return []
        else:
            return value.split(self.import_multiple_value_splitter)

    @admin.action(description="Export as XLSX")
    def export_as_xlsx(self, request, queryset):
        """
        Method to write data queryset into an xlsx file.
        """
        meta = self.model._meta

        if self.export_fields:
            field_names = self.export_fields
        else:
            field_names = [field.name for field in meta.fields]

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename={}.xlsx".format(meta)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append([f.split("__")[0] for f in field_names])

        for obj in queryset:
            row = []

            for field_name in field_names:
                field_parts = field_name.split("__")

                value = obj

                for field_part in field_parts:
                    # for a many-relation we can join the related model's field
                    # value with the splitter
                    if type(value).__name__ == "ManyRelatedManager":
                        value = self.import_multiple_value_splitter.join(
                            [getattr(r, field_part) for r in value.all()]
                        )
                    else:
                        value = getattr(value, field_part)

                # Handle ImageField specifically
                if hasattr(value, "url") and "ImageField" in str(type(value)):
                    if value:
                        value = value.url
                    else:
                        value = ""  # or some default placeholder text

                row.append(value)

            worksheet.append(row)

        workbook.save(response)

        return response

    # export_as_xlsx.short_description = "Export as XLSX"


class RecordImportError(Exception):
    errors: dict[str, list[int]]

    def __init__(self, errors):
        self.errors = errors

        super().__init__()


def custom_titled_filter(title):
    class Wrapper(admin.FieldListFilter):
        def __new__(cls, *args, **kwargs):
            instance = admin.FieldListFilter.create(*args, **kwargs)
            instance.title = title
            return instance

    return Wrapper
